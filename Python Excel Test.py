#SheetChange Automator
#Data Source: https://www.ers.usda.gov/data-products/food-price-outlook/food-price-outlook/#Consumer%20Price%20Index
#missing data, validating data, duplicate
import xlsxwriter
import pandas as pd
from openpyxl import load_workbook

# Load the Excel files into dataframes
df1 = pd.read_excel('apples_2013.xlsx')
df2 = pd.read_excel('apples_2020.xlsx')

df1.drop(index=df1.index[-1],axis=0,inplace=True)
df2.drop(index=df2.index[-1],axis=0,inplace=True)

#Print 2013 apples data
print('Dataframe for 2013 apple prices')
print(df1.head())        # View the first few rows
print(df1.info())        # Get information about columns and data types

#Print 2020 apples data
print('Dataframe for 2020 apple prices')
print(df2.head())        # View the first few rows
print(df2.info())        # Get information about columns and data types

# Create a new Excel workbook
workbook = xlsxwriter.Workbook('Apple Price Index.xlsx')

with pd.ExcelWriter("Apple Price Index.xlsx") as writer:
    df1.to_excel(writer, sheet_name='2013 Apple Prices', index=False)
    df2.to_excel(writer, sheet_name="2020 Apple Prices", index=False)

#this added the dataframe to the new excel and closed the workbook

# Specify the path to your existing Excel workbook
workbook_path = 'Apple Price Index.xlsx'

# Load the existing workbook for editing
workbook = load_workbook(workbook_path)

#create Differences worksheet
worksheet3 = workbook.create_sheet('Differences')

# Extract the 'ColumnToCopy' from the DataFrame
Form = df1['Form']
Retail_2013_Price = df1['Average retail price ']
Retail_2020_Price = df2['Average retail price ']

# Create a new DataFrame for the target worksheet
df3 = pd.DataFrame()

# Add the extracted column to the new DataFrame
df3['Form'] = Form
df3['2013 Average Retail Price'] = Retail_2013_Price
df3['2020 Average Retail Price'] = Retail_2020_Price
df3['Price Difference'] = df3['2020 Average Retail Price'] - df3['2013 Average Retail Price']

print('Output of df3')
print(df3)

# Save the new DataFrame to the target worksheet
with pd.ExcelWriter('Apple Price Index.xlsx', engine='openpyxl', mode='a') as writer:
    df3.to_excel(writer, sheet_name='Differences', index=False)
